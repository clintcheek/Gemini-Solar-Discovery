from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import DocumentRecord

HEADERS = [
    "Solar Classification", "Solar Score", "Matched Keywords", "Qualification Source",
    "Document Type", "Recorded Date", "Document Number", "Grantor / Debtor",
    "Grantee / Secured Party", "Town", "Legal Description", "Book / Volume / Page",
    "Search Source", "Source URL", "County", "Status", "Property Address", "Loan Amount",
]
WIDTHS = [22, 12, 38, 22, 28, 16, 24, 38, 38, 22, 60, 24, 28, 48, 14, 14, 46, 16]


def _values(record: DocumentRecord) -> list[object]:
    return [
        record.solar_classification, record.solar_score, record.matched_keywords,
        record.qualification_source, record.document_type, record.loan_date,
        record.document_number, record.grantor, record.grantee, record.town,
        record.legal_description, record.book_volume_page, record.search_term,
        record.source_url, record.county, record.status, record.property_address,
        float(record.loan_amount) if record.loan_amount is not None else None,
    ]


def _format_sheet(sheet, rows: list[DocumentRecord], table_name: str) -> None:
    sheet.append(HEADERS)
    for record in rows:
        sheet.append(_values(record))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet["R"][1:]:
        cell.number_format = "$#,##0.00"
    if rows:
        table = Table(displayName=table_name, ref=f"A1:R{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)


def export_excel(records: Iterable[DocumentRecord], outdir: Path) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    path = outdir / f"Dallas_Solar_Discovery_Build_0005_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    rows = list(records)
    qualified = [r for r in rows if r.solar_score >= 40]
    review = [r for r in rows if 20 <= r.solar_score < 40]

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Run Summary"
    summary.append(["Metric", "Value"])
    summary.append(["All acquired filings", len(rows)])
    summary.append(["Qualified solar filings", len(qualified)])
    summary.append(["Review queue", len(review)])
    summary.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 24

    inventory = workbook.create_sheet("Acquisition Inventory")
    _format_sheet(inventory, rows, "AcquisitionInventory")
    solar = workbook.create_sheet("Qualified Solar")
    _format_sheet(solar, qualified, "QualifiedSolar")
    review_sheet = workbook.create_sheet("Review Queue")
    _format_sheet(review_sheet, review, "SolarReviewQueue")

    workbook.save(path)
    if not path.exists() or path.stat().st_size == 0:
        raise RuntimeError("Workbook export verification failed")
    return path
